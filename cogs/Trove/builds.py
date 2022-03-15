# Priority: 1
import asyncio
import os
import string
from datetime import datetime

import discord
from discord.ext import commands
from openpyxl import Workbook, utils
from openpyxl.styles import Alignment, Font, PatternFill
from utils.buttons import BuildsPickView, Confirm, GemBuildsView, Paginator
from utils.CustomObjects import CEmbed
from utils.objects import BuildType, GameClass, Values


class Builds(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.values = Values()
        self.bot.Trove.last_updated = 1647025641

    @commands.command(slash_command=True, help="Show gear for a class")
    @commands.cooldown(1, 120, commands.BucketType.user)
    @commands.bot_has_permissions(embed_links=1)
    async def gear(
        self,
        ctx,
        _class: GameClass=commands.Option(name="class", default=None, description="Input class name"),
        *,
        build_type: BuildType=commands.Option(name="build_type", default=None, description="Input build type [Farm | DPS | Health]")
    ):
        await ctx.defer()
        if build_type and not _class:
            return await ctx.reply("Class is not valid", ephemeral=True, delete_after=10)
        page = None
        all_gears = self.get_all_gear_pages(ctx)
        if build_type:
            if not self.bot.Trove.values.gear_builds[_class.name][build_type]["enabled"]:
                build_type = None
            else:
                try:
                    page = all_gears[_class.name][build_type]
                except:
                    pass
        view = BuildsPickView(ctx, _class, build_type, all_gears)
        if not _class:
            content = f"Pick a **class** to get builds for."
        elif not build_type:
            content = f"Pick a build type for **{_class}**"        
        else:
            content = None
        view.message = await ctx.reply(content, embed=page["embed"] if page else None, view=view)

    @commands.command(aliases=["gu"], hidden=True)
    async def gear_update(self, ctx):
        if ctx.author.id not in [565097923025567755,237634733264207872]:
            return
        self.values.update_gear_builds()
        await ctx.reply("Updated gear builds.")

    @commands.command(aliases=["gem", "gems"], help="Show gem builds for a class.")
    @commands.cooldown(1, 180, commands.BucketType.user)
    @commands.bot_has_permissions(embed_links=1)
    async def build(self, ctx, build_id=commands.Option(default=None, description="Load your's or someone's saved build")):
        if ctx.author.id == 565097923025567755:
            ctx.command.reset_cooldown(ctx)
        # if ctx.author.id not in [565097923025567755]:
        #     return await ctx.send("This command is under update...please wait a few minutes.", delete_after=15)
        build = None
        build_data = None
        if build_id:
            builds_data = await self.bot.db.db_users.find_one({"builds.saved": {"$elemMatch": {"code": build_id}}}, {"builds.saved": 1})
            if builds_data:
                for build_data in builds_data["builds"]["saved"]:
                    if build_data["code"] != build_id:
                        continue
                    if build_data["creator"] != ctx.author.id and not build_data["public"]:
                        return await ctx.reply(f"This build is not public.")
                    build = build_data["config"]
                    break
                build_data["views"] += 1
                await self.bot.db.db_users.update_one({"_id": build_data["creator"], "builds.saved.code": build_id}, {f"$inc": {"builds.saved.$.views": 1}})
                creator = await self.bot.try_user(build_data['creator'])
            else:
                await ctx.reply(f"Build ID doesn't correspond to any build in database.\nThis command takes no arguments like class or type, it's just `{ctx.prefix}build`", delete_after=15, ephemeral=True)
        view = GemBuildsView(ctx, build_data=build_data)
        view.message = await ctx.reply(
            content="Builds will only be calculated once all **Required** fields are filled in." if not build else f"Loaded **{build_data['code']}** by {creator.mention}",
            view=view,
            allowed_mentions=discord.AllowedMentions.none()    
        )

    @commands.group(slash_command=True, aliases=["build_list", "builds_list"], help="Manage and show saved builds")
    async def builds(self, ctx):
        ...

    @builds.command(slash_command=True, name="list", help="List of all your saved builds.")
    async def _build_list_list(self, ctx):
        data = await self.bot.db.db_users.find_one({"_id": ctx.author.id}, {"builds": 1})
        if not data["builds"]["saved"]:
            return await ctx.reply("You have no saved builds.")
        pages = await self._build_list_pages(data["builds"]["saved"])
        view = Paginator(ctx, pages, start_end=True)
        view.message = await ctx.reply(embed=pages[0]["embed"], view=view)

    @builds.command(slash_command=True, name="public", help="List of all public builds.")
    async def _build_list_public(self, ctx):
        data = await self.bot.db.db_users.find({"builds.saved": {"$elemMatch": {"public": True}}}).distinct("builds.saved")
        if not data:
            return await ctx.reply("No public builds available.")
        data = [b for b in data if b["public"]]
        pages = await self._build_list_pages(data, True)
        view = Paginator(ctx, pages, start_end=True)
        view.message = await ctx.reply(embed=pages[0]["embed"], view=view)

    @builds.command(slash_command=True, name="like", aliases=["upvote"], help="Like a public build.")
    async def _build_list_like(self, ctx, build_id=commands.Option(description="Build ID to like.")):
        try:
            build = await self._get_build(build_id, public=True)
        except Exception as e:
            return await ctx.reply(str(e), ephemeral=True)
        if build is None:
            return await ctx.send("You can't like a private build.")
        if build["creator"] == ctx.author.id:
            return await ctx.reply("You cannot like your own builds.", ephemeral=True)
        if ctx.author.id in build["likes"]:
            update = ('disliked', 'pull')
        else:
            update = ('liked', 'push')
        await self.bot.db.db_users.update_one({"_id": build["creator"], "builds.saved.code": build_id}, {f"${update[1]}": {"builds.saved.$.likes": ctx.author.id}})
        return await ctx.send(f"You {update[0]} build with ID **{build_id}**")
    
    @builds.command(slash_command=True, name="delete", help="Delete one of your saved builds.")
    async def _build_list_delete(self, ctx, build_id=commands.Option(description="Build ID to edit.")):
        try:
            build = await self._get_build(build_id, own=ctx.author)
        except Exception as e:
            return await ctx.reply(str(e), ephemeral=True)
        view = Confirm(ctx, timeout=30)
        view.message = await ctx.send(f"Are you sure you want to delete build with ID **{build['code']}**", view=view, delete_after=30)
        await view.wait()
        try:
            await view.message.delete()
        except:
            pass
        if view.value is None:
            return await ctx.send("Time out! Build deletion was cancelled.", delete_after=10, ephemeral=True)
        elif not view.value:
            return await ctx.send("Build deletion was cancelled.", delete_after=10, ephemeral=True)
        await self.bot.db.db_users.update_one({"_id": ctx.author.id}, {"$pull": {"builds.saved": build}})
        return await ctx.reply(f"Build with id **{build['code']}** was deleted.", ephemeral=True)

    @builds.group(slash_command=True, name="set", aliases=["edit"], help="Edit your build's Metadata.")
    async def _build_list_set_values(self, ctx):
        ...

    @_build_list_set_values.command(slash_command=True, name="name", help="Edit your build's name.")
    async def _build_list_set_name(self, ctx, 
        build_id=commands.Option(description="Build ID to edit."),
        *, name=commands.Option(description="Change command's name. Max: 32 Characters")):
        if len(name) > 32:
            return await ctx.reply("Maximum of 32 characters on name.", ephemeral=True)
        try:
            build = await self._get_build(build_id, own=ctx.author)
        except Exception as e:
            return await ctx.reply(str(e), ephemeral=True)
        await self.bot.db.db_users.update_one({"_id": build["creator"], "builds.saved.code": build_id}, {f"$set": {"builds.saved.$.name": name}})
        return await ctx.reply(f"Build name is now:\n```\n{name}\n```", ephemeral=True)

    @_build_list_set_values.command(slash_command=True, name="description", aliases=["about"], help="Edit your build's description.")
    async def _build_list_set_description(self, ctx, 
        build_id=commands.Option(description="Build ID to edit."), 
        *, description=commands.Option(description="Change command's description. Max: 180 Characters")):
        if len(description) > 180:
            return await ctx.reply("Maximum of 180 characters on description.", ephemeral=True)
        try:
            build = await self._get_build(build_id, own=ctx.author)
        except Exception as e:
            return await ctx.reply(str(e), ephemeral=True)
        await self.bot.db.db_users.update_one({"_id": build["creator"], "builds.saved.code": build_id}, {f"$set": {"builds.saved.$.description": description}})
        return await ctx.reply(f"Build description is now:\n```\n{description}\n```", ephemeral=True)

    @_build_list_set_values.command(slash_command=True, name="public", aliases=["private"], help="Toggle build public state.")
    async def _build_list_set_public(self, ctx, build_id=commands.Option(description="Build ID to edit.")):
        try:
            build = await self._get_build(build_id, own=ctx.author)
        except Exception as e:
            return await ctx.reply(str(e), ephemeral=True)
        await self.bot.db.db_users.update_one({"_id": build["creator"], "builds.saved.code": build_id}, {f"$set": {"builds.saved.$.public": not build["public"]}})
        if not build["public"]:
            return await ctx.reply("Build is now 🌐 Public", ephemeral=True)
        else:
            return await ctx.reply("Build is now 🔒 Private", ephemeral=True)

    @commands.Cog.listener()
    async def on_raw_reaction_add(self, payload):
        if self.bot.get_user(payload.user_id).bot:
            return
        if str(payload.emoji) != "👁️":
            return
        channel = self.bot.get_channel(payload.channel_id)
        try:
            msg = await channel.fetch_message(payload.message_id)
        except:
            return
        if msg.author.id == self.bot.user.id and msg.flags.suppress_embeds:
            content = msg.content
            try:
                await msg.edit(content=None, view=None, suppress=False)
                await msg.clear_reaction("👁️")
            except:
                pass
            await asyncio.sleep(300)
            try:
                await msg.edit(content=content, view=None, suppress=True)
                await msg.add_reaction("👁️")
            except:
                pass

    @commands.command(hidden=True)
    async def ea(self, ctx):
        stats = []
        for ally in self.bot.Trove.values.allies:
            for stat, _ in ally.stats.items():
                if stat not in stats:
                    stats.append(stat)
        stats.sort()
        letters = string.ascii_uppercase
        stats = {stats[i]: letters[i+1] for i in range(len(stats))}
        def adjust_width(ws):
            def is_merged_horizontally(cell):
                cell_coor = cell.coordinate
                if cell_coor not in ws.merged_cells:
                    return False
                for rng in ws.merged_cells.ranges:
                    if cell_coor in rng and len(list(rng.cols)) > 1:
                        return True
                return False

            for col_number, col in enumerate(ws.columns, start=1):
                col_letter = utils.get_column_letter(col_number)

                max_length = max(len(str(cell.value or "")) for cell in col if not is_merged_horizontally(cell))
                adjusted_width = (max_length + 2) * 1
                ws.column_dimensions[col_letter].width = adjusted_width
        wb = Workbook()
        wb.remove_sheet(wb.active)
        ws = wb.create_sheet(title="Allies")
        ws["A1"] = "Name"
        for stat, collumn in stats.items():
            ws[f"{collumn}1"] = stat
        allies = sorted(self.bot.Trove.values.allies, key=lambda x: x.name.lower())
        fill = PatternFill(start_color='00000000',
                            end_color='00000000',
                            fill_type='solid')
        font = Font(color='00747474')
        for i in range(len(allies)):
            ally = allies[i]
            row = i + 2
            ws[f"A{row}"] = ally.name
            for stat, value in ally.stats.items():
                collumn = stats[stat]
                ws[f"{collumn}{row}"] = str(value)
            for i in range(len(stats.keys())+1):
                collumn = letters[i]
                ws[f"{collumn}{row}"].fill = fill
                ws[f"{collumn}{row}"].font = font
                ws[f"{collumn}{row}"].alignment = Alignment(horizontal='center')
        adjust_width(ws)
        for i in range(len(allies)):
            row = i + 2
            ally = discord.utils.get(allies, name=ws[f"A{row}"].value)
            ws[f"A{row}"] = f'=HYPERLINK("{ally.url}", "{ally.name}")'
        _file = "cache/allies.xlsx"
        wb.save(_file)
        await ctx.send(file=discord.File(_file))
        os.remove(_file)

    async def _build_list_pages(self, builds, public=False):
        pages = []
        builds.sort(key=lambda x: (-len(x["likes"]), -x["views"], -x["last_updated"]))
        paged_builds = self.bot.utils.chunks(builds, 4)
        i = 0
        for paged_build in paged_builds:
            i += 1
            e = CEmbed()
            e.set_author(name=("Public List" if public else "Private List") + f" [{i}/{len(paged_builds)}]")
            for build in paged_build:
                name = f"[{build['code']}] {build['name']}" if build['name'] else f"{build['code']}"
                if not public:
                    name = ("🌐 " if build["public"] else "🔒 ") + name
                if public:
                    user = await self.bot.try_user(build["creator"])
                    name += f" by {user}"
                value = f"Created on <t:{build['created_at']}:D>\n"
                value += f"Updated on <t:{build['last_updated']}:D>\n"
                metrics = []
                if build["public"]:
                    metrics.append(f"❤️ {len(build['likes'])}")
                metrics.append(f"👁️ {build['views']}\n")
                value += " | ".join(metrics)
                value += f"```\n{build['description']}\n```"
                e.add_field(name=name, value=value, inline=False)
            page = {
                "content": None,
                "page": i,
                "embed": e
            }
            pages.append(page)
        return pages

    async def _get_build(self, build_id, own=None, public=False):
        query = {
            "builds.saved.code":  build_id
        }
        if own:
            query["_id"] = own.id
        builds_data = await self.bot.db.db_users.find(query, {"builds": 1}).distinct("builds.saved")
        if not builds_data:
            raise Exception(("You don't have a" if own else "There's no") + " build with that ID")
        for build in builds_data:
            if build_id == build["code"]:
                if public and not build["public"]:
                    continue
                return build

    def readable_stats(self, stats, num=True, fixed=False, or_split=False):
        text = ""
        if not or_split:
            text += "```\n"
            if not fixed:
                for i in range(len(stats)):
                    if num:
                        text += f"{i+1}. "
                    else:
                        text += "● "
                    stat_text = stats[i]
                    # if stat_text == 'Explosive Epilogue':
                    #     stat_text == 'Explosive Epi.'
                    # if stat_text == "Freerange Electrolytic Crystals":
                    #     stat_text = "Freerange Elec. Cryst."
                    # stat_text.replace(" Vial", "")
                    text += f"{stat_text}\n"
            else:
                text += "/".join(["● "+i for i in stats])
            text += "```"
        else:
            text += "```css\n"
            text += "\n#or\n".join(["● "+i for i in stats])
            text += "\n```"

        return text

    def get_all_gear_pages(self, ctx, ):
        all_gears = {}
        for _class in self.bot.Trove.values.classes:
            if _class.name not in all_gears.keys():
                all_gears[_class.name] = {}
            for k, v in self.get_gear_page(ctx, _class).items():
                all_gears[_class.name][k] = v
        return all_gears

    def get_gear_page(self, ctx, _class):
        build_types = ["light", "farm", "tank"]
        builds = {}
        for build_type in build_types:
            class_build = self.values.gear_builds[_class.name][build_type]
            if not class_build["enabled"]:
                continue
            if build_type == "light":
                build_type = "dps"
                build = "DPS"
            elif build_type == "tank":
                build = "Tanking"
            if build_type == "farm":
                build = "Farming"
            trovesaurus = f"https://trovesaurus.com/builds/{_class.name.replace(' ', '%20')}/{build_type}"
            #tiers = ["<:Star5:841015868930523176>", "<:Star4:841015868863283240>", "<:Star3:841015868716220447>", "<:Star2:841015868854501376>", "<:Star1:841015868993175592>"]
            e = CEmbed(description=f"**Check on [Trovesaurus]({trovesaurus})**",color=discord.Color.random(), timestamp=datetime.utcfromtimestamp(self.bot.Trove.last_updated))
            e.description += f"\n\nRating: " + '<:Star:841018551087530024>' * (6-class_build['tier']) + '<:StarOutline:841018551418880010>' * (class_build['tier'] - 1)
            e.set_author(name=f"In depth {build_type} build for {_class.name}", icon_url=_class.image)
            e.add_field(name="<:hat:834512699585069086>Hat", value=self.readable_stats(class_build["hat"]))
            e.add_field(name="<:sword:834512699593064518>Weapon", value=self.readable_stats(class_build["weapon"]))
            e.add_field(name="<:face:834512699441938442>Face", value=self.readable_stats(class_build["face"]))
            e.add_field(name="<:banner:834512699391475742>Banner", value=self.readable_stats(class_build["banner"], False))
            e.add_field(name="<:ring:923960128401719337>Ring", value=self.readable_stats(class_build["ring"], False))
            e.add_field(name="<:food:834512699424505886>Food", value=self.readable_stats(class_build["food"], False))
            e.add_field(name="<:flask:834512699479687228>Flask", value=self.readable_stats(class_build["flask"], False, or_split=True))
            e.add_field(name="<:emblem:834512699332755536>Emblems", value=self.readable_stats(class_build["emblem"], False))
            e.add_field(name="<:qubesly:834512699361853530>Ally", value=self.readable_stats(class_build["ally"], False, True))
            e.add_field(name="<:blue_emp_gem:834506349433585735>Gem Abilities", value=self.readable_stats(class_build["gems"], False))
            e.add_field(name="<:subclass:834512699576025119>Subclass", value=self.readable_stats(class_build["subclass"], False, or_split=True))
            e.add_field(name="\u200b", value="Note: These builds are subjective, contact Sly#0511 for suggestions.\nThanks to **Nikstar** for putting these builds together.", inline=False)
            e.set_footer(text=f"For gem builds use {ctx.prefix}build {_class.short} {build_type} | Last updated on")
            if build_type == "dps":
                build_type = "light"
            builds[build_type.lower()] = {
                "description": f"Build aimed at improving {build}",
                "embed": e,
                "link": trovesaurus
            }
        return builds

def setup(bot):
    bot.add_cog(Builds(bot))
