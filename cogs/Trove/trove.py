# Priority: 1
import asyncio
import functools
import io
import re
import typing
import urllib.request as urlget
from datetime import datetime, timedelta

import aiohttp
import discord

import requests as getre
from bs4 import BeautifulSoup
from discord.ext import commands, tasks
from openpyxl import load_workbook
from PIL import Image

import utils.checks as perms  # pyright: reportMissingImports=false
from utils.objects import MemeType, TroveTime, Values
from utils.buttons import Paginator


class Trove(commands.Cog):
#--------------- Module Management --------------#
    def __init__(self, bot):
        self.bot = bot
        self.luxion_inventory = {}
        self.session = aiohttp.ClientSession()
        self.bot.loop.create_task(self.cog_setup())
        self.sheet = None
        self.memento_list = None
        self.values = Values()
        self.trovetime = TroveTime()
        self.sheet_timer.start()

    def cog_unload(self):
        self.sheet_timer.cancel()
        self.bot.loop.create_task(self.cog_close())

    async def cog_close(self):
        await self.session.close()

    async def cog_setup(self):
        self.classes = [["SH", "Shadow Hunter"], ["IS", "Ice Sage"], ["DL", "Dracolyte"], ["CB", "Candy Barbarian"], ["RV", "Revenant"], ["CM", "Chloromancer"],
                        ["FT", "Fae Trickster"], ["TR", "Tomb Raiser"], ["BR", "Boomeranger"], ["KT", "Knight"], ["VG", "Vanguardian"], ["DT", "Dino Tamer"],
                        ["LL", "Lunar Lancer"], ["PC", "Pirate Captain"], ["NN", "Neon Ninja"], ["GS", "Gunslinger"]]
        db_bot = (await self.bot.db.db_bot.find_one({"_id": "0511"}))["mastery"]
        self.max_live_mastery = db_bot["max_live_mastery"]
        self.max_pts_mastery = db_bot["max_pts_mastery"]
        self.max_console_mastery = db_bot["max_console_mastery"]
        self.max_live_mastery_geode = db_bot["max_live_mastery_geode"]
        self.max_pts_mastery_geode = db_bot["max_pts_mastery_geode"]
        self.max_console_mastery_geode = db_bot["max_console_mastery_geode"]

#--------------- Tasks --------------#

    @tasks.loop(seconds=60)
    async def sheet_timer(self):
        def load_sheets():
            urlget.urlretrieve("https://docs.google.com/spreadsheets/d/1hsz9Xhf52xjX0pcfb95Tm3-MvVh9lMfq4bGi5fo7zCs/export?format=xlsx&id=1hsz9Xhf52xjX0pcfb95Tm3-MvVh9lMfq4bGi5fo7zCs", 'data/sheets/luxion_sheet.xlsx')
            try:
                self.sheet = load_workbook(filename="data/sheets/luxion_sheet.xlsx")
            except:
                pass
            urlget.urlretrieve("https://docs.google.com/spreadsheets/d/1YBf3__CPCy9iL4HDEoF1_q88vaFtAR6mA4GZxIzOEG8/export?format=xlsx&id=1YBf3__CPCy9iL4HDEoF1_q88vaFtAR6mA4GZxIzOEG8", 'data/sheets/memento_list_sheet.xlsx')
            self.memento_list = load_workbook(filename="data/sheets/memento_list_sheet.xlsx", data_only=True)
        await self.bot.loop.run_in_executor(None, load_sheets)

#--------------- Events----------------#

    @commands.Cog.listener()
    async def on_message(self, message):
        if not message.guild:
            return
        ctx = await self.bot.get_context(message)
        if ctx.channel.id == 812762716908552272 and not ctx.valid and not ctx.author.id == self.bot.user.id:
            return await message.delete()
        if ctx.channel.id == 859433409016496128 and not ctx.author.bot:
            await ctx.send("Use command retard...", delete_after=10)
            return await message.delete()
        TT_role = message.guild.get_role(722182894394802206)
        if TT_role and TT_role in message.role_mentions and message.channel.id in [721801555174621234] and TT_role in message.author.roles:
            already = []
            e = discord.Embed(description=f"React to join in.\n", color=self.bot.progress)
            e.set_author(name=f"{message.author.display_name} wants to start a delve run.", icon_url=message.author.avatar)
            msg = await message.channel.send(embed=e)
            reactions = [str(self.bot.get_emoji(722924965535416442)), str(self.bot.get_emoji(722924930173239418)), str(self.bot.get_emoji(722924913261805579)), str(self.bot.get_emoji(722924946602328165))]
            for reaction in reactions:
                await msg.add_reaction(reaction)
            while True:
                def check(reaction, user):
                    return reaction.message.id == msg.id and str(reaction) in reactions and TT_role in user.roles and user.id not in already
                reaction, user = await self.bot.wait_for("reaction_add", check=check)
                already.append(user.id)
                the_class = None
                if str(reaction) == str(self.bot.get_emoji(722924965535416442)):
                    the_class = "Gunslinger"
                if str(reaction) == str(self.bot.get_emoji(722924930173239418)):
                    the_class = "Ice Sage"
                if str(reaction) == str(self.bot.get_emoji(722924913261805579)):
                    the_class = "Candy Barbarian"
                if str(reaction) == str(self.bot.get_emoji(722924946602328165)):
                    the_class = "Dracolyte"
                e.description = e.description + f"\n{user.mention} --> **{the_class}**"
                await msg.edit(embed=e)
                try:
                    await msg.remove_reaction(reaction, user)
                except:
                    pass
                if len(already) == 8:
                    break
            description = e.description.replace("React to join in.", "Team was assembled.")
            e = discord.Embed(description=description, color=self.bot.success)
            e.set_author(name=f"{message.author.display_name} assembled a delve run team.", icon_url=message.author.avatar)
            await msg.edit(embed=e)
        post_channel = self.bot.get_channel(706947765187641426)
        bug_channel = 706939493130829915
        if message.channel.id != bug_channel:
            return
        bug_content = {}
        for line in message.content.split("\n"):
            if line.lower().startswith("server:"):
                content = line.casefold().replace("server:", "")
                if content.startswith(" "):
                    content = content[1:]
                bug_content["server"] = content
            if line.lower().startswith("time:"):
                content = line.casefold().replace("time:", "")
                if content.startswith(" "):
                    content = content[1:]
                bug_content["time"] = content
            if line.lower().startswith("bug:"):
                content = line.casefold().replace("bug:", "")
                if content.startswith(" "):
                    content = content[1:]
                bug_content["bug"] = content
            if line.lower().startswith("expected:"):
                content = line.casefold().replace("expected:", "")
                if content.startswith(" "):
                    content = content[1:]
                bug_content["expected"] = content
            if line.lower().startswith("media:"):
                content = line.replace("Media:", "").replace("media:", "").replace("MEDIA:", "")
                if content.startswith(" "):
                    content = content[1:]
                bug_content["media"] = content
        if len(bug_content.keys()) < 3:
            try:
                await message.delete()
            except:
                pass
            return
        try:
            e=discord.Embed(description=bug_content["bug"], color=self.bot.comment, timestamp=datetime.utcnow())
            e.set_author(name=message.author.display_name)
            e.add_field(name="Server", value=bug_content["server"], inline=False)
            e.add_field(name="Time", value=bug_content["time"], inline=False)
            if "expected" in bug_content:
                e.add_field(name="Expected", value=bug_content["expected"],inline=False)
            if "media" in bug_content:
                e.add_field(name="Media", value=bug_content["media"], inline=False)
            if message.attachments:
                e.set_image(url=message.attachments[0].url)
            e.set_footer(text="Sent in UTC at")
        except:
            await message.delete()
            await message.channel.send(f"Invalid format {message.author.mention}", delete_after=10)
            return
        await post_channel.send(embed=e)
        await message.add_reaction("✅")

#--------------- Commands --------------#

    def primes(self, start, end):
        for num in range(2, end+1):
            if num < 2:
                continue
            kill = False
            for n in range(2,int(num/2)+1):
                if num%n==0:
                    kill = True
            if kill:
                continue
            yield num

    @commands.command(slash_command=True, help="Get the calculated rewards for paragon levels")
    async def paragon_rewards(self, ctx,
        maximum: int=commands.Option(name="maximum", description="Maximum level of the range."),
        minimum: int=commands.Option(name="minimum", default=1, description="Minimum level of the range.")):
        if not 0 < maximum <= 1000:
            return await ctx.reply("Maximum must be greater than 0 and smaller or equal to 1000.")
        if not 0 < minimum <= 999:
            return await ctx.reply("Maximum must be greater than 0 and smaller or equal to 999.")
        if not minimum < maximum:
            return await ctx.reply("Maximum must be greater than minimum.")
        primes = list(self.primes(minimum, maximum))
        trovian_loops = 0
        primal_loops = 0
        paragon_pinatas = 0
        for i in range(minimum, maximum):
            trovian_loops += 1
            if i in primes:
                primal_loops += 1
                paragon_pinatas += 1
        await ctx.send(f"**Start Level:** {minimum}\n**End Level:** {maximum}\n**Trovian Loops:** {trovian_loops}\n**Primal Loops:** {primal_loops}\n**Paragon Pinatas:** {paragon_pinatas} ({paragon_pinatas*2} if patron)\n**Pinata Rolls:** {paragon_pinatas}")

    @commands.command(slash_command=True, message_command=False, help="Search for an ally name or stats/abilities")
    async def search_ally(self, ctx,
        name=commands.Option(name="name", default=None, description="Search for an ally through name"),
        stat=commands.Option(name="stat", default=None, description="Search for an ally through stat"),
        ability=commands.Option(name="ability", default=None, description="Search for an ally through ability")):
        allies = self.bot.Trove.values.allies
        if name or stat or ability:
            query = {
                "name": name.lower() if name else None,
                "stat": stat.lower() if stat else None,
                "ability": ability.lower() if ability else None
            }
            def query_filter(item: tuple):
                def loop_filter(stuff: list, sep: str, equals=False):
                    value = False
                    for thing in stuff:
                        if not equals:
                            value = sep in thing.lower()
                        else:
                            value = thing.lower() == sep
                        if value:
                            break
                    return value
                return (
                    (item.name.lower().startswith(query["name"]) if query["name"] else True) and
                    (loop_filter(item.stats, query["stat"], True) if query["stat"] else True) and
                    (loop_filter(item.abilities, query["ability"]) if query["ability"] else True)
                )
            allies = list(filter(query_filter, allies))
            if name:
                allies.sort(key=lambda x: x.name)
            elif stat:
                def sorter(item: tuple, negative=False):
                    for stat, raw_value in item.stats.items():
                        value = raw_value
                        if stat.lower() != query["stat"]:
                            continue
                        if isinstance(value, str):
                            value = value.replace("%", "")
                        if stat.lower() in ["incoming damage"]:
                            value = -(float(value))
                        else:
                            value = float(value)
                        if isinstance(raw_value, str) and "%" in raw_value:
                            value = value * 1000
                        return (value, len(item.abilities))
                allies.sort(key=lambda x: sorter(x), reverse=True) 
        pages = []
        paged_results = self.bot.utils.chunks(allies, 6)
        for raw_page in paged_results:
            e = discord.Embed()
            e.set_author(name=f"Page {paged_results.index(raw_page)+1} of {len(paged_results)}")
            e.description = ""
            e.color = discord.Color.random()
            for ally in raw_page:
                e.description += f"[**{ally.name}**]({ally.url})" + ("\nStats: " + " | ".join([f"**{stat}={value}**" for stat, value in ally.stats.items()]) if ally.stats else "") + ("\nAbilities: " + "\n".join(ally.abilities) if ally.abilities else "") + "\n\n"
            page = {
                "page": paged_results.index(raw_page),
                "content": None,
                "embed": e
            }
            pages.append(page)
        if pages:
            view = Paginator(ctx, pages, start_end=True)
            view.message = await ctx.send(embed=pages[0]["embed"], view=view)
        else:
            await ctx.send(f"No allies found. Use `{ctx.prefix}feedback <my feedback>` to send a missing or wrong ally.")

    @commands.command(aliases=["lightsteps", "lsteps"])
    async def light_steps(self, ctx):
        e = discord.Embed(title="Light Steps", color=discord.Color.random())
        e.set_image(url="https://i.imgur.com/FTN3hcc.png")
        e.set_footer(text=f"Data provided by {self.bot.get_user(523714260090748938)}")
        await ctx.send(embed=e)

    @commands.command(aliases=["chatcmd"])
    async def chats(self, ctx):
        chats = {
            "levi": "Join **Leviathan** farms in Geode Topside.",
            "delve": "Join **Delve** farms in Delves.",
            "trade": "Trade your goodies or currency with other players.",
            "fivefarm": "Join **5 Star Dungeon** farms in Geode Topside.",
            "dragon": "Join **Dragon Fragment** farms.",
            "nitro": "Join **Nitro Glitterine** farms in Geode Topside.",
            "ganda": "Join **Nitro Glitterine** farms in Geode Topside.",
            "egg": "Join **Egg** farms. __(During **Bunfest** event only)__"
        }
        e = discord.Embed()
        e.color = discord.Color.random()
        e.description = ""
        for chat, description in sorted(chats.items()):
            e.description += f"`/join {chat}` -> {description}\n"
        e.set_author(name="Joinable Chats in game", icon_url=self.bot.user.avatar)
        await ctx.send(embed=e)

 # Other Commands

    @commands.command(aliases=["effort"])
    async def effort_leaderboard(self, ctx):
        e = discord.Embed()
        e.set_author(name="Effort Contest Objectives", icon_url=ctx.guild.me.avatar)
        e.description = """```py\n#Dungeon Objectives:
    1 star dungeon objective – 1 point
    3 star dungeon objective – 3 points
    5 star dungeon objective – 5 points if completed on time, 1 point if completed after time ran out
    Leviathan lair – 10 points
#Hourly Challenge Objectives:
    Rampage – each tier 8 points
    Coin Collection – each tier 8 points
    Biome Challenge – each tier 8 points
#Delves:
    Public Delve floor completed – 15 points
    Private Delve floor completed – 20 points
    Challenge Delve floor completed – 25 points
#Adventures:
    Trovian Adventures – 3 points
    Luminopolis Adventures – 3 points
    Club Adventures – 3 points
    Geodian Adventures – 3 points
#Bomber Royale:
    Bomber Royale kills – 1 point every few enemies defeated
    Bomber Royale Win – 5 points
#Geode Caves:
    Reliquary Filled – 5 points
    Companion Egg found – 5 points
#Gardening:
    Watering Plants – 1 point every few dozen watered
    Harvesting Plants – 1 point every few dozen harvested
#General Activities:
    Defeating Monsters – 1 point every few dozen enemies defeated
    Defeating World Bosses – 1 point every few enemies defeated
    Destroying Blocks - 1 point every few hundred blocks destroyed```Credit: **Avish233**"""
        await ctx.send(embed=e)

    @commands.command(name="betterdiscord", aliases=["bbd"])
    async def _better_discord(self, ctx):
        e = discord.Embed(description="Go to Custom CSS tab in your settings and then paste the following line onto it.```css\n@import url('https://slynx.xyz/trovegametags');```**Hit update and save.**", color=discord.Color.random())
        e.set_author(name="Better Discord Trove Tags", icon_url=self.bot.user.avatar)
        await ctx.send(embed=e)

    @commands.command(slash_command=True, help="Display server time and next reset", aliases=["time", "stime"])
    async def server_time(self, ctx):
        today = self.trovetime.now
        time = today.isoformat().split('T')[1][:5]
        reset = int((today + timedelta(days=1, hours=-today.hour+11, minutes=-today.minute, seconds=-today.second, microseconds=today.microsecond)).timestamp())
        await ctx.send(embed=discord.Embed(description=f"🕐 Server Time {time}\nNext reset <t:{reset}:R> <t:{reset}:F>", color=self.bot.comment))

    @commands.command(slash_command=True, help="Display maximum Trove Mastery in PC Trove.", aliases=["m"])
    async def mastery(
        self,
        ctx,
        mode: typing.Literal["live", "pts", "console"]=commands.Option(name="server", default=None, description="Pick a server."),
        update=commands.Option(name="value", default=None, description="Amount to update mastery with.")
    ):
        if mode == None:
            e=discord.Embed(color=self.bot.comment, timestamp=datetime.utcfromtimestamp((await self.bot.db.db_bot.find_one({"_id": "0511"}))["mastery"]["mastery_update"]))
            e.set_author(name="Trove Max Mastery", icon_url="https://i.imgur.com/t5aCX3u.png")
            live_level, live_points, live_level_points = self.bot.utils.points_to_mr(self.max_live_mastery)
            pts_level, pts_points, pts_level_points = self.bot.utils.points_to_mr(self.max_pts_mastery)
            console_level, console_points, console_level_points = self.bot.utils.points_to_mr(self.max_console_mastery)
            live_text = f"\nLevel: **{live_level}**"
            pts_text = f"\nLevel: **{pts_level}**"
            console_text = f"\nLevel: **{console_level}**"
            if live_points > 0:
                live_text += f"\nExtra Points: **{live_points}**"
            if pts_points > 0:
                pts_text += f"\nExtra Points: **{pts_points}**"
            if console_points > 0:
                console_text += f"\nExtra Points: **{console_points}**"
            e.add_field(name="Live Maximum", value=f"Total Points: **{self.max_live_mastery}**{live_text}\nMax PR: **{self.max_pr(live_level - 500)}**" + f"\n{self.bot.utils.progress_bar(live_level_points, live_points, 20)}")
            e.add_field(name="PTS Maximum", value=f"Total Points: **{self.max_pts_mastery}**{pts_text}\nMax PR: **{self.max_pr(pts_level - 500, True)}**" + f"\n{self.bot.utils.progress_bar(pts_level_points, pts_points, 20)}")
            e.add_field(name="Console Maximum", value=f"Total Points: **{self.max_console_mastery}**{console_text}\nMax PR: **{self.max_pr(console_level - 500, False, True)}**" + f"\n{self.bot.utils.progress_bar(console_level_points, console_points, 20)}")
            e.set_footer(text="Last updated on")
            await ctx.send(embed=e)
            return
        if ctx.author.id in [562659821476773942] and mode == "console":
            ...
        elif ctx.author.id in [237634733264207872, 565097923025567755, 209712946534809600]:
            ...
        else:
            return
        if mode.lower() == "live":
            db = "mastery.max_live_mastery"
        elif mode.lower() == "pts":
            db = "mastery.max_pts_mastery"
        elif mode.lower() == "console":
            db = "mastery.max_console_mastery"
        else:
            await ctx.send("Invalid server!")
            return
        try:
            value = int(update)
        except:
            return await ctx.send("Invalid value!")
        if update.isdigit():
            await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$set": {db: value, "mastery.mastery_update": datetime.utcnow().timestamp()}})
            await ctx.send(f"**{mode.capitalize()}** Max Mastery was set to **{value}**", ephemeral=True)
        else:
            if value > 0:
                await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$inc": {db: value}, "$set":{"mastery.mastery_update": datetime.utcnow().timestamp()}})
                await ctx.send(f"Added **{value}** to **{mode.capitalize()} Max Mastery**", ephemeral=True)
            else:
                await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$inc": {db: value}, "$set":{"mastery.mastery_update": datetime.utcnow().timestamp()}})
                await ctx.send(f"Removed **{value}** from **{mode.capitalize()} Max Mastery**", ephemeral=True)
        db_bot = await self.bot.db.db_bot.find_one({"_id": "0511"})
        self.max_live_mastery = db_bot["mastery"]["max_live_mastery"]
        self.max_console_mastery = db_bot["mastery"]["max_console_mastery"]
        #await self._update_trovesaurus_mr(self.max_live_mastery)
        difference = self.max_live_mastery - db_bot["mastery"]["max_pts_mastery"]
        if difference > 0 and mode.lower() == "live":
            await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$inc": {"mastery.max_pts_mastery": difference}})
            db_bot = await self.bot.db.db_bot.find_one({"_id": "0511"})
        self.max_pts_mastery = db_bot["mastery"]["max_pts_mastery"]
    
    async def _update_trovesaurus_mr(self, name, mastery):
        payload = {
            "SetMaxMastery": "",
            "Key": self.bot.keys["Trovesaurus"]["MasteryToken"],
            "Score": str(mastery),
            "Name": name
        }
        await self.bot.AIOSession.post("https://trovesaurus.com/", data=payload)

    @commands.command(slash_command=True, help="Display maximum Geode Mastery in PC Trove.", aliases=["mg", "gm"])
    async def mastery_geode(
        self,
        ctx,
        mode: typing.Literal["live", "pts", "console"]=commands.Option(name="server", default=None, description="Pick a server."),
        update=commands.Option(name="value", default=None, description="Amount to update mastery with.")
    ):
        if mode == None:
            e=discord.Embed(color=self.bot.comment, timestamp=datetime.utcfromtimestamp((await self.bot.db.db_bot.find_one({"_id": "0511"}))["mastery"]["geode_mastery_update"]))
            e.set_author(name="Geode Max Mastery", icon_url="https://i.imgur.com/mJRDFtT.png")
            live_level, live_points, live_level_points = self.bot.utils.points_to_mr(self.max_live_mastery_geode)
            pts_level, pts_points, pts_level_points = self.bot.utils.points_to_mr(self.max_pts_mastery_geode)
            console_level, console_points, console_level_points = self.bot.utils.points_to_mr(self.max_pts_mastery_geode)
            live_text = f"\nLevel: 100 ||(**{live_level}**)||"
            pts_text = f"\nLevel: 100 ||(**{pts_level}**)||"
            console_text = f"\nLevel: 100 ||(**{console_level}**)||"
            if live_points > 0:
                live_text += f"\nExtra Points: **{live_points}**"
            if pts_points > 0:
                pts_text += f"\nExtra Points: **{pts_points}**"
            if console_points > 0:
                console_text += f"\nExtra Points: **{console_points}**"
            e.add_field(name="Live Maximum", value=f"Total Points: **{self.max_live_mastery_geode}**{live_text}" + f"\n{self.bot.utils.progress_bar(live_level_points, live_points, 20)}")
            e.add_field(name="PTS Maximum", value=f"Total Points: **{self.max_pts_mastery_geode}**{pts_text}" + f"\n{self.bot.utils.progress_bar(pts_level_points, pts_points, 20)}")
            e.add_field(name="Console Maximum", value=f"Total Points: **{self.max_console_mastery_geode}**{console_text}" + f"\n{self.bot.utils.progress_bar(console_level_points, console_points, 20)}")
            e.set_footer(text="Last updated on")
            await ctx.send(embed=e)
            return
        if ctx.author.id in [562659821476773942] and mode == "console":
            ...
        elif ctx.author.id in [237634733264207872, 565097923025567755, 209712946534809600]:
            ...
        else:
            return
        if mode.lower() == "live":
            db = "mastery.max_live_mastery_geode"
        elif mode.lower() == "pts":
            db = "mastery.max_pts_mastery_geode"
        elif mode.lower() == "console":
            db = "mastery.max_console_mastery_geode"
        else:
            await ctx.send("Invalid server!")
            return
        if update.lower().startswith("+"):
            try:
                value = int(update)
            except:
                await ctx.send("Invalid value!")
                return
            await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$inc": {db: value}, "$set":{"mastery.geode_mastery_update": datetime.utcnow().timestamp()}})
            await ctx.send(f"Added **{value}** to **{mode.capitalize()} Max Geode Mastery**", delete_after=8, ephemeral=True)
        elif update.lower().startswith("-"):
            try:
                value = int(update)
            except:
                await ctx.send("Invalid value!")
                return
            await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$inc": {db: value}, "$set":{"mastery.geode_mastery_update": datetime.utcnow().timestamp()}})
            await ctx.send(f"Removed **{value}** from **{mode.capitalize()} Max Geode Mastery**", delete_after=8,  ephemeral=True)
        else:
            try:
                value = int(update)
            except:
                await ctx.send("Invalid value!")
                return
            await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$set": {db: value, "mastery.geode_mastery_update": datetime.utcnow().timestamp()}})
            await ctx.send(f"**{mode.capitalize()}** Max Geode Mastery was set to **{value}**", delete_after=8,  ephemeral=True)
        db_bot = await self.bot.db.db_bot.find_one({"_id": "0511"})
        self.max_live_mastery_geode = db_bot["mastery"]["max_live_mastery_geode"]
        self.max_console_mastery_geode = db_bot["mastery"]["max_console_mastery_geode"]
        if self.max_live_mastery_geode > db_bot["mastery"]["max_pts_mastery_geode"]:
            await self.bot.db.db_bot.update_one({"_id": "0511"}, {"$inc": {"mastery.max_pts_mastery_geode": self.max_live_mastery_geode - db_bot["mastery"]["max_pts_mastery_geode"]}})
            db_bot = await self.bot.db.db_bot.find_one({"_id": "0511"})
        self.max_pts_mastery_geode = db_bot["mastery"]["max_pts_mastery_geode"]

    @commands.command(slash_command=True, help="Display maximum Magic Find in Trove PC.", aliases=["mf"])
    async def magic_find(self, ctx):
        e=discord.Embed(color=self.bot.comment, timestamp=datetime.utcfromtimestamp((await self.bot.db.db_bot.find_one({"_id": "0511"}))["mastery"]["mastery_update"]))
        e.set_author(name="Max Magic Find", icon_url="https://i.imgur.com/sCIbgLX.png")
        live_level, _, _ = self.bot.utils.points_to_mr(self.max_live_mastery)
        pts_level, _, _ = self.bot.utils.points_to_mr(self.max_pts_mastery)
        live_mf = self.values.base_mf + (live_level - 500)
        pts_mf = self.values.base_mf + (pts_level - 500)
        def mastery_table(mf):
            return f"""
```
  Magic Find  | Normal | Patron |
--------------|--------|--------|
  Normal      |  {round(mf*1.5)}  |  {round(mf*2*1.5)} |
+ Sunday      |  {round((mf+100)*1.5)}  |  {round((mf+200)*2*1.5)} |
+ Clov & Elix |  {round((mf+100+75+50)*1.5)}  |  {round((mf+200+75+50)*2*1.5)} |
```
"""
        e.add_field(name="Live Max Magic Find", value=mastery_table(live_mf), inline=False)
        e.add_field(name="PTS Max Magic Find", value=mastery_table(pts_mf))
        e.set_footer(text="Last updated on")
        await ctx.send(embed=e)

    @commands.command(slash_command=True, help="Shows people currently playing Trove in discord. [Depends on activity]", aliases=["pltrove"])
    @commands.cooldown(1, 180, commands.BucketType.guild)
    async def playingtrove(self, ctx, show_all: typing.Literal["True"]=commands.Option(name="show_all", default=None, description="Whether to show from all servers or not.")):
        if not show_all:
            ctx.command.reset_cooldown(ctx)
        msg = await ctx.send("<a:loading:844624767239192576> Looking for players...")
        def get_members():
            checked = []
            playingtrove = []
            fake = []
            for member in self.bot.get_all_members() if show_all else ctx.guild.members:
                if member.id in checked:
                    continue
                checked.append(member.id)
                for activity in member.activities:
                    if activity.name == "Trove":
                        if "application_id" not in dir(activity):
                            continue
                                                            #       PC                   XBOX
                        if activity.application_id not in [363409000399634432, 438122941302046720]:
                            fake.append([member.id, member.guild.name, member.guild.id])
                            continue
                        if activity.application_id == 363409000399634432:
                            platform = "<:windows:839445248517079050>"
                        if activity.application_id == 438122941302046720:
                            platform = "<:xbox:839439645359603713>"
                        if ctx.guild.get_member(member.id):
                            is_in = True
                        else:
                            is_in = False
                        playingtrove.append([str(member), member.id, activity.created_at.timestamp(), platform, is_in])
                        break
            return checked, playingtrove, fake
        task = functools.partial(get_members)
        checked, playingtrove, fake = await self.bot.loop.run_in_executor(None, task)
        playingtrove.sort(key=lambda x: x[2])
        pages = self.bot.utils.chunks(playingtrove, 10)
        page = 0
        e = discord.Embed(color=discord.Color.random())
        e.set_author(name=f"Playing Trove... ({len(playingtrove)})", icon_url="https://i.imgur.com/sCIbgLX.png")
        if len(pages) == 0:
            e.description = "There's no one in this server playing trove right now!"
            return await msg.edit(content=None, embed=e)
        e.set_footer(text="This depends on Discord Game Activity.\n⭐-> Is in current server.")
        def check(reaction, user):
            return user == ctx.author and reaction.message.id == msg.id and str(reaction.emoji) in ["◀️", "▶️"]
        await msg.add_reaction("◀️")
        await msg.add_reaction("▶️")
        while True:
            stime = datetime.utcnow().timestamp()
            e.description = ""
            for member, _, time, platform, is_in in pages[page]:
                e.description += f"{platform}`{member}` playing for {self.bot.utils.time_str(stime - time)[1]}{' ⭐' if is_in else ''}\n"
            try:
                await msg.edit(content=None, embed=e)
                reaction, _ = await self.bot.wait_for("reaction_add", check=check, timeout=60)
            except:
                try:
                    await msg.clear_reactions()
                except:
                    try:
                        await msg.remove_reaction("◀️")
                        await msg.remove_reaction("▶️")
                    except:
                        break
                    break
                break
            if str(reaction.emoji) == "◀️":
                page -= 1
                if page < 0:
                    page = len(pages) - 1
            if str(reaction.emoji) == "▶️":
                page += 1
                if page > len(pages) - 1:
                    page = 0
            try:
                await reaction.remove(ctx.author)
            except:
                pass
        
    @commands.command(slash_command=True, help="Shows when luxion is around and it's inventory.", aliases=["lux"])
    async def luxion(self, ctx):
        if self.trovetime.is_luxion:            
            initial_id = 84
            initial_id += int(self.trovetime.luxion_time[1])
            if not self.luxion_inventory.get(str(initial_id)):
                msg = await ctx.send("<a:loading:844624767239192576> Looking for inventory in <https://trovesaurus.com/luxion>", delete_after=5)
                await asyncio.sleep(3)
                self.luxion_inventory[str(initial_id)] = []
                trovesaurus = await (await self.session.get(f"https://trovesaurus.com/luxion/visit/{initial_id}")).text()
                soup = BeautifulSoup(trovesaurus, "html.parser")
                data = soup.findAll("div", {"class": "navbar-secondary border"})
                if not data:
                    try:
                        await msg.delete()
                    except:
                        ...
                    return await ctx.send(f"Luxion is available. It will be gone {self.trovetime.luxion_end_rwts('R')}, yet inventory couldn't be fetched, try again later.")
                items = []
                for d in data:
                    d = d.parent
                    item = {}
                    item["name"] = d.find("a").getText().strip()
                    item["link"] = d.find("a")["href"]
                    item["price"] = int(d.find("p").getText().strip())
                    limit = d.find("span", {"class": "text-muted"})
                    if limit:
                        text = limit.getText()
                        limit = int(re.findall(r"([0-9]+)", text)[0])
                    item["limit"] = str(limit)
                    item_page = await (await self.session.get(item["link"])).text()
                    soup = BeautifulSoup(item_page, "html.parser")
                    #item["tradeable"] = -bool(soup.findAll("strong", text="Cannot be traded"))
                    items.append(item)
                self.luxion_inventory[str(initial_id)] = items
            e = discord.Embed(description=f"Luxion is available. It will be gone {self.trovetime.luxion_end_rwts('R')}\nItem's info was fetched from **[Trovesaurus](https://trovesaurus.com/luxion)** made by **Etaew**", color=self.bot.comment)
            e.set_author(name="Current Luxion's Inventory", icon_url="https://i.imgur.com/9eOV0JD.png")
            for item in self.luxion_inventory[str(initial_id)]:
                text = ""
                text += f'**[{item["name"]}]({item["link"]})**'
                text += f"\nPrice: **{item['price']} <:dragon_coin:858061506074378276>**"
                #text += f"\nStatus: **{'Tradeable' if item['tradeable'] else 'Not Tradeable'}**"
                text += f"\nLimit: **{item['limit']}**"
                e.add_field(name="\u200b", value=text)
            e.add_field(name="\u200b", value="\u200b")
            e.set_footer(text="Inventory from")
            e.timestamp = self.trovetime.rw_time(self.trovetime.luxion_start)
            await ctx.send(embed=e)
        else:
            e = discord.Embed(
                description=f"It will be available on **{self.trovetime.luxion_start_rwts('F')}** until **{self.trovetime.luxion_end_rwts('F')}** **{self.trovetime.luxion_start_rwts('R')}**",
                color=self.bot.comment)
            e.set_author(name="Luxion is not available.", icon_url="https://i.imgur.com/0N0PYdp.png")
            await ctx.send(embed=e)

    @commands.command(slash_command=True, help="Shows when corruxion is around.", aliases=["nlux"])
    async def corruxion(self, ctx):
        if self.trovetime.is_corruxion:
            e = discord.Embed(description=f"Corruxion is available. It will be gone {self.trovetime.corruxion_end_rwts('R')}", color=self.bot.comment)
            e.description += "\n\nCorruxion items have a global purchase limit of 10 per visit."
            items = {
                "Empowered Water Gem Box": [15, "https://trovesaurus.com/item/lootbox/gems/empowered_blue"],
                "Empowered Fire Gem Box": [15, "https://trovesaurus.com/item/lootbox/gems/empowered_red"],
                "Empowered Air Gem Box": [15, "https://trovesaurus.com/item/lootbox/gems/empowered_yellow"],
                "Empowered Cosmic Gem Box": [15, "https://trovesaurus.com/item/lootbox/gems/empowered_cosmic"],
                "Lustrous Gem Box": [25, "https://trovesaurus.com/item/lootbox/gems/foci"],
                "Double Experience Potion": [3, "https://trovesaurus.com/item/consumable/xp_double"],
                "Bound Brilliance": [2, "https://trovesaurus.com/item/crafting/boundbrilliance"],
                "Chaos Dragon Egg Fragment": [1, "https://trovesaurus.com/item/dragon/egg/chaos_notrade_fragment"],
            }
            for item, price in items.items():
                text = f'[**{item}**]({price[1]})'
                text += f"\nPrice: **{price[0]} <:dragon_coin:858061506074378276>**"
                e.add_field(name="\u200b", value=text)
            e.add_field(name="\u200b", value="\u200b")
            e.set_author(name="Corruxion", icon_url="https://i.imgur.com/BPNdE1w.png")
            await ctx.send(embed=e)
        else:
            e = discord.Embed(
                description=f"It will be available on **{self.trovetime.corruxion_start_rwts('F')}** until **{self.trovetime.corruxion_end_rwts('F')}** **{self.trovetime.corruxion_start_rwts('R')}**",
                color=self.bot.comment)
            e.set_author(name="Corruxion is not available.", icon_url="https://i.imgur.com/BPNdE1w.png")
            await ctx.send(embed=e)

    @commands.command(
        slash_command=True,
        help="Calculates coefficient given certain values."
    )
    async def sigil(
        self,
        ctx,
        powerrank: int=commands.Option(name="power_rank", description="Input power rank for sigil. [0-40000]"),
        masteryrank: int=commands.Option(name="mastery_rank", description="Input mastery rank for sigil. [0-849]")
    ):
        if powerrank > 40000:
            await ctx.send("Power rank can't be that high. Max: 40000")
            return
        if masteryrank > 849:
            await ctx.send("Mastery rank can't be that high. Max: 849")
            return
        sigil = self.get_sigil(powerrank, masteryrank)
        if not sigil:
            await ctx.send("That sigil is not in my database.")
            return
        e = discord.Embed(color=self.bot.comment)
        e.set_image(url=f"attachment://sigil.png")
        await ctx.send(file=discord.File(sigil, filename="sigil.png"), embed=e)

    @commands.command(slash_command=True, help="Shows current weekly [May be slow to update]", aliases=["week"])
    async def weekly(self, ctx):
        if not self.sheet:
            await ctx.send("Try again in a few seconds, fetching info...")
            return
        sheet = self.sheet["Chaos Chests + Weekly"]
        now = datetime.utcnow() - timedelta(hours=11)
        if sheet["A10"].value.timestamp() + 86400 * 6 < now.timestamp():
            await ctx.send("Info fetched is outdated, try again later.")
            return
        num = 10
        if sheet["A10"].value.timestamp() > now.timestamp():
            num = 11
        e = discord.Embed(description="\nWeekly's info was fetched from **[this spreadsheet](https://trove.summerhaas.com/luxion)** made by **__SummerHaas__**", color=self.bot.comment, timestamp=sheet[f"A{num}"].value)
        e.set_author(name="Chaos Chests & Weekly")
        e.add_field(name="CC Loot", value=sheet[f"B{num}"].value if sheet[f"B{num}"].value else "Unknown")
        e.add_field(name="Weekly bonus", value=sheet[f"C{num}"].value)
        e.add_field(name="\u200b", value="\u200b")
        e.add_field(name="Weekly Deal", value=sheet[f"D{num}"].value)
        if sheet[f"E{num}"].value:
            e.add_field(name="Event", value=sheet[f"E{num}"].value)
        else:
            e.add_field(name="\u200b", value="\u200b")
        e.add_field(name="\u200b", value="\u200b")
        await ctx.send(embed=e)

    @commands.command(
        slash_command=True,
        help="Shows list of mementos and their depths in delves.",
        name="memento_list",
        aliases=["ml"]
    )
    async def _memento_list(self, ctx):#, _filter: typing.Union[MemeType, str]=commands.Option(name="filter", default=None, description="Input a memento, boss or biome to filter by memento type.")):
        return await ctx.send("https://slynx.xyz/trove/depths/")
        if not self.memento_list:
            await ctx.send("Try again in a few seconds, fetching info...")
            return
        ml = self.memento_list["This Week"]
        depth_column = ml["B"]
        header = "🟩 **Creature**\n🟦 **Boss**\n🟪 **Biome**\n⚔️ **Mount Bosses**\n\n"
        depths_listed = []
        body = ""
        for cell in depth_column:
            val = ml[f'C{cell.row}'].value
            if not cell.value:
                break
            if val:
                if _filter and _filter not in ["biome", "boss", "creature"]:
                    if _filter.lower() not in val.replace(' (Biome)', '').replace(' (Boss)', '').replace(' (Creature)', '').lower():
                        continue
                elif _filter and _filter in ["biome", "boss", "creature"]:
                    if _filter not in val.lower():
                        continue
                if " (Creature)" in val:
                    val = '🟩 **' + val.replace(' (Creature)', '') + "**"
                if " (Boss)" in val:
                    val = '🟦 **' + val.replace(' (Boss)', '') + "**"
                if " (Biome)" in val:
                    val = '🟪 **' + val.replace(' (Biome)', '') + "**"
                body += f"`{cell.value} -> `{val}"
                #merch = ml[f'D{cell.row}'].value
                #biome = ml[f'E{cell.row}'].internal_value
                #if merch:
                #    body += f" 💰"
                if cell.row != 1:
                    boss = ml[f'A{cell.row}'].value
                    if boss:
                        body += f" ⚔️ {boss}"
                #if biome:
                #    body += "\n"
                #    body += f"`       `🌌{biome}"
                body += "\n\n"
                depths_listed.append(int(cell.value))
        body_texts = body.split("\n")
        def get_text(texts):
            textys = []
            text = ""
            i = 0
            for texty in texts:
                if i == 30:
                    i = 0
                    textys.append(text)
                    text = ""
                text += texty + "\n"
                i += 1
            if text != "":
                textys.append(text)
            return textys
        textys = get_text(body_texts)
        if _filter:
            textys = ["No mementos found matching that filter."] if textys[0] == "\n" else textys
        if not _filter:
            missing = "\nMissing Depths: "
            for i in range(110, depths_listed[-1] + 1):
                if i not in depths_listed:
                    missing += f"`{i}` "
            if missing == "\nMissing Depths: ":
                missing == ""
        else:
            missing = ""
        page = 0
        e = discord.Embed(
            title="Thanks to Alpha Legacy Staff for keeping this updated through the spreadsheet",
            url="https://docs.google.com/spreadsheets/d/1YBf3__CPCy9iL4HDEoF1_q88vaFtAR6mA4GZxIzOEG8", 
            description=f"**Check Web version [here](https://slynx.xyz/trove/depths{'/' + _filter.replace(' ', '%20') if _filter else ''})**\n\n" + (header if not _filter else "") + textys[page] + missing, color=self.bot.comment)
        e.set_author(name="Challenge Depth Memento List", icon_url=self.bot.user.avatar)
        sent = await ctx.send(embed=e)
        if len(textys) == 1:
            return
        await sent.add_reaction("◀️")
        await sent.add_reaction("▶️")
        def check(reaction, user):
            return user == ctx.author and reaction.message.id == sent.id and str(reaction.emoji) in ["◀️", "▶️"]
        while True:
            try:
                reaction, user = await self.bot.wait_for("reaction_add", check=check, timeout=60)
            except:
                try:
                    await sent.clear_reactions()
                except:
                    break
                break
            if str(reaction.emoji) == "◀️":
                page -= 1
                if page < 0:
                    page = len(textys) - 1
            if str(reaction.emoji) == "▶️":
                page += 1
                if page > len(textys) - 1:
                    page = 0
            e.description = header + textys[page] + missing
            try:
                await sent.edit(embed=e)
            except:
                pass
            try:
                await reaction.remove(ctx.author)
            except:
                pass

    @commands.command(
        slash_command=True,
        help="Shows list of depths in delves.",
        name="depth_list",
        aliases=["dl"]
    )
    async def _depth_list(self, ctx):#, *, _filter=commands.Option(name="filter", default=None, description="Input a memento, boss or biome name to filter.")):
        return await ctx.send("https://slynx.xyz/trove/depths/")
        if not self.memento_list:
            await ctx.send("Try again in a few seconds, fetching info...")
            return
        ml = self.memento_list["This Week"]
        header = "Memento Types:\n🟩 **Creature**\n🟦 **Boss**\n🟪 **Biome**\n⬛ **Unknown**\n\n"
        depths = []
        for i in range(141):
            memento = ml[f'C{i+1}'].value
            boss = ml[f'D{i+1}'].value
            biome = ml[f'E{i+1}'].value
            if not memento and not boss and (not biome or biome == "#N/A"):
                continue
            else:
                if memento:
                    regex = r"(?i)([a-z0-9\.]+(?: )?[a-z0-9\.]+(?: )?[a-z0-9\.]+(?: )?[a-z0-9\.]+)(?: )?\((creature|boss|biome)\)"
                    res = re.findall(regex, memento)
                depth = {
                    "level": 110+i,
                    "memento": res[0][0] if memento else "Unknown",
                    "memento_type": res[0][1].lower() if memento else None,
                    "boss": boss if boss else "Unknown",
                    "biome": biome if biome else "Unknown"
                }
                if _filter and _filter not in " ".join([memento.lower() if memento else "", boss.lower() if boss else "", biome.lower() if biome and biome != "#N/A" else ""]):
                    continue
                depths.append(depth)
        depths.sort(key=lambda x: x["level"])
        missing = ""
        if not _filter:
            missing = [f"`{i}`" for i in range(110, depths[-1]["level"]+1) if i not in [l["level"] for l in depths]]
            if missing:
                missing = "Missing Depths: "+ " ".join(missing)
        def chunks(lst, n):
            result = []
            for i in range(0, len(lst), n):
                result.append(lst[i:i + n])
            return result
        depths = chunks(depths, 15)
        page = 0
        pretext = f"Thanks to Alpha Legacy Staff for keeping this updated through this [spreadsheet](https://docs.google.com/spreadsheets/d/1YBf3__CPCy9iL4HDEoF1_q88vaFtAR6mA4GZxIzOEG8)\n\n**Check Web version [here](https://slynx.xyz/trove/depths{'/' + _filter.replace(' ', '%20') if _filter else ''})**\n\n" +header + (missing if missing else "")
        e = discord.Embed(description=pretext, color=self.bot.comment)
        e.set_author(name="Challenge Depth List", icon_url=self.bot.user.avatar)
        e.set_footer(text="🏆 -> Memento | 👑 -> Boss | ⛰️ -> Biome")
        sent = None
        emojis = {
            "creature": "🟩",
            "boss": "🟦",
            "biome": "🟪"
        }
        while True:
            e.clear_fields()
            if not len(depths) and _filter:
                e.add_field(name="\u200b", value="No results matching filter were found.", inline=False)
            else:
                for depth in depths[page]:
                    memento = depth['memento'] 
                    e.add_field(name=f"{emojis[depth['memento_type']] if depth['memento_type'] else '⬛'} Depth {depth['level']}", value=f"🏆`{memento}`\n👑`{depth['boss']}`\n⛰️`{depth['biome']}`") #{emojis[depth['memento_type']] if memento != 'Unknown' else ''}
                if len(depths[page]) in [5, 8, 11, 14, 17]:
                    e.add_field(name="\u200b", value="\u200b")
            if not sent:
                sent = await ctx.send(embed=e)
                if len(depths) > 1:
                    await sent.add_reaction("◀️")
                    await sent.add_reaction("▶️")
                    def check(reaction, user):
                        return user == ctx.author and reaction.message.id == sent.id and str(reaction.emoji) in ["◀️", "▶️"]
                else:
                    return
            else:
                await sent.edit(embed=e)
            try:
                reaction, _ = await self.bot.wait_for("reaction_add", check=check, timeout=60)
            except:
                try:
                    await sent.clear_reactions()
                except:
                    break
                break
            if str(reaction.emoji) == "◀️":
                page -= 1
                if page < 0:
                    page = len(depths) - 1
            if str(reaction.emoji) == "▶️":
                page += 1
                if page > len(depths) - 1:
                    page = 0
            try:
                await reaction.remove(ctx.author)
            except:
                pass

 # Settings Commands

    @commands.group(name="settings", aliases=["set"])
    @perms.has_permissions("administrator")
    async def _settings(self, ctx):
        ...

    @_settings.command()
    async def game_server(self, ctx):
        data = await self.bot.db.db_servers.find_one({"_id": ctx.guild.id}, {"PTS mode": 1})
        await self.bot.db.db_servers.update_one({"_id": ctx.guild.id}, {"$set": {"PTS mode": not data["PTS mode"]}})
        await ctx.reply(f"PTS Server mode set to **{not data['PTS mode']}**")

#--------------- Functions --------------#

    def get_sigil(self, pr, totalmr):
        url_mr = ""
        url_pr = ""
        # Mastery
        if 1 <= totalmr <= 9:
            url_mr = "https://i.imgur.com/ytgFQJq.png"
        if 10 <= totalmr <= 29:
            url_mr = "https://i.imgur.com/qFlsroS.png"
        if 30 <= totalmr <= 49:
            url_mr = "https://i.imgur.com/VcnmxYG.png"
        if 50 <= totalmr <= 74:
            url_mr = "https://i.imgur.com/k1uSwAJ.png"
        if 75 <= totalmr <= 99:
            url_mr = "https://i.imgur.com/ST8NDsO.png"
        if 100 <= totalmr <= 149:
            url_mr = "https://i.imgur.com/uWDUKxk.png"
        if 150 <= totalmr <= 199:
            url_mr = "https://i.imgur.com/atceNdE.png"
        if 200 <= totalmr <= 249:
            url_mr = "https://i.imgur.com/WuPD8Vi.png"
        if 250 <= totalmr <= 299:
            url_mr = "https://i.imgur.com/IG4Uncx.png"
        if 300 <= totalmr <= 349:
            url_mr = "https://i.imgur.com/222rH1k.png"
        if 350 <= totalmr <= 399:
            url_mr = "https://i.imgur.com/9hiZXCZ.png"
        if 400 <= totalmr <= 449:
            url_mr = "https://i.imgur.com/ZEAVWDA.png"
        if 450 <= totalmr <= 499:
            url_mr = "https://i.imgur.com/Z7g60mW.png"
        if 500 <= totalmr <= 549:
            url_mr = "https://i.imgur.com/BzJj2lL.png"
        if 550 <= totalmr <= 599:
            url_mr = "https://i.imgur.com/n04blVH.png"
        if 600 <= totalmr <= 649:
            url_mr = "https://i.imgur.com/EufwyxA.png"
        if 650 <= totalmr <= 699:
            url_mr = "https://i.imgur.com/wq5D6Iw.png"
        if 700 <= totalmr <= 749:
            url_mr = "https://i.imgur.com/sYTJNHU.png"
        if 750 <= totalmr <= 799:
            url_mr = "https://i.imgur.com/jpx5SZE.png"
        if 800 <= totalmr <= 849:
            url_mr = "https://i.imgur.com/6kS2qH3.png"
        # Power
        if 1 <= pr <= 249:
            url_pr = "https://i.imgur.com/ytgFQJq.png"
        if 250 <= pr <= 549:
            url_pr = "https://i.imgur.com/Lu25aXz.png"
        if 550 <= pr <= 899:
            url_pr = "https://i.imgur.com/BLgFldK.png"
        if 900 <= pr <= 1199:
            url_pr = "https://i.imgur.com/93uU1oA.png"
        if 1200 <= pr <= 2499:
            url_pr = "https://i.imgur.com/AOWt4On.png"
        if 2500 <= pr <= 4999:
            url_pr = "https://i.imgur.com/l6TejUh.png"
        if 5000 <= pr <= 7499:
            url_pr = "https://i.imgur.com/rF0ezqj.png"
        if 7500 <= pr <= 9999:
            url_pr = "https://i.imgur.com/G27zi8j.png"
        if 10000 <= pr <= 14999:
            url_pr = "https://i.imgur.com/cTCm7QK.png"
        if 15000 <= pr <= 19999:
            url_pr = "https://i.imgur.com/9QN4F1W.png"
        if 20000 <= pr <= 24999:
            url_pr = "https://i.imgur.com/FpXg2LU.png"
        if 25000 <= pr <= 29999:
            url_pr = "https://i.imgur.com/E7mqm6F.png"
        if 30000 <= pr <= 34999:
            url_pr = "https://i.imgur.com/x7EqPKj.png"
        if 35000 <= pr <= 39999:
            url_pr = "https://i.imgur.com/cdT2Hub.png"
        if 40000 <= pr:
            url_pr = "https://i.imgur.com/6m6jfYq.png"
        if url_mr == "" or url_pr == "":
            return False
        image_mr = getre.get(url_mr).content
        img1 = Image.open(io.BytesIO(image_mr))
        image_pr = getre.get(url_pr).content
        img2 = Image.open(io.BytesIO(image_pr))
        img1.paste(img2, (0, 0), img2)
        data = io.BytesIO()
        img1.save(data, "PNG")
        data.seek(0)
        return data

    def select_class(self, args):
        pclass = None
        for c in self.classes:
            if c[0].lower() == args.lower():
                pclass = c[1]
        pclass = [c[1] for c in self.classes if c[0].lower() == args.lower()][-1]
        if pclass == None:
            for c in self.classes:
                if c[1].lower().startswith(args.lower()):
                    pclass = c[1]
        return pclass

    def max_pr(self, mastery, pts=False, console=False):
        sources = {
            "hat": 1698,
            "weapon": 1698,
            "face": 1698,
            "gems": 24769,
            "dragons": 1620,
            "banner": 350,
            "ring": 1513 if not console else 987,
            "ally": 75,
            "class_level": 450,
            "subclass": 90,
            "emblem": 50 * 2,
            "flask": 50,
            "mastery_geode": 5 * 100,
            "mastery_rank": 4 * 500,
            "500_plus_mastery": mastery
        }
        return sum(sources.values())

def setup(bot):
    bot.add_cog(Trove(bot))
